from openpyxl import load_workbook, Workbook

# Crear un nuevo archivo Excel
def create_excel(file_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos"
    sheet.append(["ID", "Nombre", "Edad"])  # Crear encabezados
    workbook.save(file_name)

# Leer datos de un archivo Excel
def read_excel(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row)

# Actualizar un dato en Excel
def update_excel(file_name, row, col, value):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    sheet.cell(row=row, column=col, value=value)
    workbook.save(file_name)

# Eliminar un dato en Excel (en este caso, eliminar contenido de una celda)
def delete_excel(file_name, row, col):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    sheet.cell(row=row, column=col, value=None)
    workbook.save(file_name)

# Uso de las funciones
file = "datos.xlsx"
create_excel(file)
update_excel(file, 2, 1, 1)  # Agregar un ID en la fila 2, columna 1
update_excel(file, 2, 2, "Juan")  # Agregar un nombre
update_excel(file, 2, 3, 25)  # Agregar una edad
read_excel(file)
delete_excel(file, 2, 2)  # Borrar el nombre
read_excel(file)
