# Conceptos basicos en openpyxl

# from openpyxl import Workbook

'''# Para crear un nuevo libro de excel
workbook = Workbook()
sheet = workbook.active # Selecciona la hoja activa
sheet.title = "Mi primera hoja" # Cambia el nombre de la hoja

# Guardar el archivo Excel
workbook.save("mi_primer_archivo.xlsx")
print("Archivo creado con éxito") # Para indicar por consola que se realizo la creacion '''


'''# Aprender a escribir datos en las celdas del documento xlsx

# Paso 1, crear un libro xlsx
workbook = Workbook()
sheet = workbook.active
sheet.title = "Register"

# Escribir datos en celdas especificas
# sheet.cell(row=1, column=1, value="ID") # De esta forma se escribe en la celda A1
# sheet.cell(row=1, column=2, value="Name")
# sheet.cell(row=1, column=3, value="LastName")
# sheet.cell(row=1, column=4, value="Age")

#Agregar datos de un alumno
# sheet.cell(row=2, column=1, value="1")
# sheet.cell(row=2, column=2, value="Maxito")
# sheet.cell(row=2, column=3, value="Trujillo")
# sheet.cell(row=2, column=4, value="26")


# Escribir datos por filas
sheet.append(["ID", "Name", "LastName", "Age"]) # Para la primera fila

# Agregar datos del alumno
sheet.append(["1", "Maxito", "Trujillo", "26"])
sheet.append(["2", "Gisela", "Vargas", '19'])
# Guardar el archivo
workbook.save("Table.xlsx")
print("Listo")'''

'''# 3. Como leer datos de un archivo Excel usando openpyxl

from openpyxl import load_workbook

# Cargar un archivo excel
workbook = load_workbook("Lista de alumnos.xlsx")
sheet = workbook.active # Selecciona la hoja activa

# Mostrar el titulo de la hoja activa
print(f'La hoja activa es {sheet.title}')

# Leer las filas de la hoja
values = sheet.iter_rows(values_only=True)
for row in values:
    array = []
    for i in row:
        if i != None:
            array.append(i)
    if array != []:
        print(array)
        
# Para obtener el dato de una celda especifica es
print(f'Valor del A1 {sheet.cell(row=1, column=1).value}')
'''

'''from openpyxl import load_workbook

# Cargar el archivo excel
workbook = load_workbook("Lista de alumnos.xlsx")
sheet = workbook.active

# sheet.cell(row=1, column=1, value="Holaaaaaa")

print(sheet.cell(row=1, column=1).value)


lista_Ciclo_1 = []
lista_Ciclo_2 = []

for i in sheet.iter_rows(values_only=True):
    arr = []
    for n in i:
        if n != None:
            arr.append(n)
    if len(arr)>=2:
        if arr[1] == "CICLO 01":
            arr.reverse()
            lista_Ciclo_1.append(arr)
        if arr[1] == "CICLO 02":
            arr.reverse()
            lista_Ciclo_2.append(arr)
sheet = workbook.close'''

'''for i in lista_Ciclo_1:
    print(i)
print('*'*25)
for i in lista_Ciclo_2:
    print(i)'''

'''# new_workbook = Workbook()
# new_sheet = new_workbook.active
# # new_sheet.title("Ciclo 01")

# for i in new_sheet.iter_rows(values_only=True):
#     new_sheet.append(i)

# new_workbook.save("Ciclo 01.xlsx")
# print("Se ha creado un archivo")
# print('*'*25)
# print(lista_Ciclo_1)
# print(lista_Ciclo_2)'''  # Esto esta mal xdxdxd

