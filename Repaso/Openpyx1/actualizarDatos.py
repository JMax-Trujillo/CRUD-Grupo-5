from openpyxl import load_workbook

workbook = load_workbook("Ciclo_01.xlsx")
sheet = workbook.active

# sheet.cell(row=1, column=5, value=sheet.cell(row=1, column=7).value)
# print(sheet.cell(row=1, column=5).value)
# print(sheet.cell(row=1, column=5).value)
# print(sheet.cell(row=1, column=6).value)
# print(sheet.cell(row=1, column=7).value)

# # Guardar cambios
# workbook.save("Ciclo_01_actualizado2.xlsx")

for row in sheet.iter_rows(min_row=2, max_row=33, min_col=1, max_col=1):
    for cell in row:
        cell.value = "Actualizado"
        
workbook.save("Actualizado_01.xlsx")
print("listo")