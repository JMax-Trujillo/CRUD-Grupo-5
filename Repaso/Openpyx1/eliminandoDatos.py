from openpyxl import load_workbook

workbook = load_workbook("Ciclo_01.xlsx")
sheet = workbook.active

sheet.cell(row=1, column=1, value=None) #Se supone debe funcionar, pero no
sheet.cell(row=1, column=1, value="") #Este si funciona, pero de otra forma q no me gusta
sheet.delete_rows(2) 
sheet.delete_cols(2)

workbook.save("Eliminado2.xlsx")
print("Se elimino")

