from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from dosListas import lista_Ciclo_1, lista_Ciclo_2
from ManejoDeHojasNuevas import orden
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

# Crear un nuevo libro
workbook = Workbook()
sheet = workbook.active

# Agregar un nuevo cuadro de excel
sheet.append(["ID", "LastName", "Name", "Ciclo"])

orden(sheet, lista_Ciclo_1)

# Aplicar fuentes
sheet["A1"].font = Font(name="Calibri", size=15, bold=True, color="FF0000")
# Rellenar la celda
fill = PatternFill(start_color="FFFF00", end_color="FF00FF", fill_type="solid")
sheet["A1"].fill = fill
# Definir bordes
thin_border = Border(left=Side(border_style="thin", color="000000"),
                     right=Side(border_style="thin", color="000000"),
                     top=Side(border_style="thin", color="000000"),
                     bottom=Side(border_style="thin", color="000000"))
# Aplicando bordes a la celda A1
sheet["A1"].border = thin_border
# Aplicando una alicacion centrada a los elementos
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")



workbook.save("styles.xlsx")
print("Pintado")