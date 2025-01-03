from openpyxl import Workbook
from dosListas import lista_Ciclo_1, lista_Ciclo_2

def orden(hoja, lista):
    key = 0
    hoja.append(["ID","Name", "LastName", "Ciclo"])
    for row in lista:
        num = row[0][-1]
        # name = row[1].replace(",", "").split(" ")
        # print(name)
        ancientName = row[1].replace(",", "").split(" ")
        name = []
        for i in ancientName:
            if i != '':
                name.append(i)
        lastName = f'{name[0]} {name[1]}'
        newListName = name[2:]
        sName = ''
        for i in newListName:
            sName += i + " "
        hoja.append([key+1, sName, lastName, num])
        key += 1

# crear el archivo xlsx
workbook = Workbook()
sheet = workbook.active
sheet.title = "Primero"

orden(sheet, lista_Ciclo_1)

newSheet = workbook.create_sheet("Segundo")

orden(newSheet, lista_Ciclo_2)

togetherSheet = workbook.create_sheet("Sistemas", index=0)

general = lista_Ciclo_1
general.extend(lista_Ciclo_2)

orden(togetherSheet, general)

# newSheet["A1"] = "Hola desde la nueva hoja!" # ubicando la celda puedo agregar un comentario

workbook.save("newList.xlsx")
print("Todo listo")

















'''key = 0
sheet.append(["ID","Name", "LastName", "Ciclo"])
for row in lista_Ciclo_1:
    num = row[0][-1]
    name = row[1].replace(", ", "").split(" ")
    lastName = f'{name[1]} {name[2]}'
    name.remove(1)
    name.remove(2)
    sName = ''
    for i in name:
        sName += i + " "
    sheet.append(key+1)
    sheet.append(sName)
    sheet.append(lastName)
    sheet.append(num)
    key += 1'''