from openpyxl import load_workbook

# Cargar el archivo excel
workbook = load_workbook("Lista de alumnos.xlsx")
sheet = workbook.active

lista_Ciclo_1 = []
lista_Ciclo_2 = []

for i in sheet.iter_rows(values_only=True):
    arr = []
    for n in i:
        if n != None:
            arr.append(n)
    if len(arr)>=2:
        if arr[1] == "CICLO 01":
            arr.reverse()
            lista_Ciclo_1.append(arr)
        if arr[1] == "CICLO 02":
            arr.reverse()
            lista_Ciclo_2.append(arr)
